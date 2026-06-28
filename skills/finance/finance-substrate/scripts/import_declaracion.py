#!/usr/bin/env python3
"""
Import data from ~/Dropbox/Declaracion/ into the finance-substrate ledger.

Handles all known formats:
  - Davivienda consolidated xlsx (from extract_bank_data.py)
  - Salary payments xlsx
  - DIAN exogena xlsx
  - DIAN e-invoices xlsx
  - International/national transfer CSVs (2023 email format)
  - 2023 transfer CSVs

Usage:
    python3 import_declaracion.py --year 2024
    python3 import_declaracion.py --year 2023
    python3 import_declaracion.py --year 2024 --source salary
    python3 import_declaracion.py --year 2024 --source all
"""


# --- Python version guard (PEP 604 union syntax requires >= 3.10) ---
import sys

if sys.version_info < (3, 10):
    raise SystemExit(
        f"finance-substrate requires Python >= 3.10. "
        f"Got {sys.version_info.major}.{sys.version_info.minor}. "
        f"Install via `brew install python@3.11` or `pyenv install 3.11 && pyenv local 3.11`."
    )
# --- end guard ---

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

DATA_DIR = Path.home() / ".finance-substrate"
LEDGER_FILE = DATA_DIR / "ledger" / "transactions.jsonl"
WITHHOLDINGS_FILE = DATA_DIR / "tax" / "withholdings.jsonl"
SALARY_FILE = DATA_DIR / "tax" / "salary-history.jsonl"
EXOGENA_FILE = DATA_DIR / "tax" / "exogena.jsonl"
INVOICES_FILE = DATA_DIR / "invoices" / "received" / "e-invoices.jsonl"
DECL_DIR = Path.home() / "Dropbox" / "Declaracion"


def ensure_data_dirs():
    for subdir in [
        "ledger", "tax", "tax/projections", "fx",
        "invoices/issued", "invoices/received",
    ]:
        (DATA_DIR / subdir).mkdir(parents=True, exist_ok=True)
    for f in [LEDGER_FILE, WITHHOLDINGS_FILE, SALARY_FILE, EXOGENA_FILE, INVOICES_FILE]:
        if not f.exists():
            f.touch()
    rules_file = DATA_DIR / "ledger" / "rules.json"
    if not rules_file.exists():
        rules_file.write_text("[]")


def tx_hash(date: str, amount: float, description: str, source: str) -> str:
    raw = f"{date}|{amount}|{description}|{source}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def load_existing_ids(file_path: Path) -> set:
    ids = set()
    if file_path.exists():
        with open(file_path) as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        entry = json.loads(line)
                        ids.add(entry.get("id", ""))
                    except json.JSONDecodeError:
                        pass
    return ids


def append_jsonl(file_path: Path, records: list) -> int:
    existing = load_existing_ids(file_path)
    new_count = 0
    with open(file_path, "a") as f:
        for rec in records:
            if rec.get("id") and rec["id"] not in existing:
                f.write(json.dumps(rec, ensure_ascii=False, default=str) + "\n")
                existing.add(rec["id"])
                new_count += 1
    return new_count


# ─── DAVIVIENDA CONSOLIDATED XLSX ───────────────────────────────────────

def import_davivienda_xlsx(year: int) -> dict:
    """Import from the extract_bank_data.py consolidated xlsx."""
    year_dir = DECL_DIR / str(year) / "Extractos Davivienda"
    xlsx_files = sorted(year_dir.glob("bank_extract_consolidated_*.xlsx"), reverse=True)

    if not xlsx_files:
        return {"source": "davivienda-xlsx", "error": "No consolidated xlsx found", "new": 0}

    xlsx_path = xlsx_files[0]  # Latest
    print(f"  Reading: {xlsx_path.name}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["All_Transactions"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    transactions = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        date_val = data.get("Date")
        if date_val is None:
            continue

        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val)[:10]

        amount = float(data.get("Amount", 0) or 0)
        description = str(data.get("Description", "")).strip()
        tid = tx_hash(date_str, amount, description, "davivienda")

        tx = {
            "id": tid,
            "date": date_str,
            "amount": amount,
            "currency": "COP",
            "description": description,
            "category": _categorize_davivienda(description, amount),
            "bank": "davivienda",
            "account": "davivienda-savings",
            "raw": {
                "document": str(data.get("Document", "")),
                "office": str(data.get("Office", "")),
                "tx_type": str(data.get("Transaction_Type", "")),
                "source_month": str(data.get("Month", "")),
            },
        }
        transactions.append(tx)

    new_count = append_jsonl(LEDGER_FILE, transactions)
    return {"source": "davivienda-xlsx", "file": xlsx_path.name, "total": len(transactions), "new": new_count}


def _categorize_davivienda(desc: str, amount: float) -> str | None:
    d = desc.upper()
    if "TRANSF INTERNACIONAL" in d or "BANCOMEX" in d:
        return "salary" if amount > 0 else "transfer-out"
    if "RENDIMIENTO" in d:
        return "interest"
    if "TRANSFERENCIA" in d or "TRANSF" in d:
        return "transfer-in" if amount > 0 else "transfer-out"
    if "COMPRA" in d:
        return "shopping"
    if "COBRO" in d or "CUOTA" in d:
        return "fees"
    if "BOLSILLO" in d:
        return "transfer-out" if amount < 0 else "transfer-in"
    if "PAGO TARJETA" in d or "PAGO TC" in d:
        return "fees"
    if "RETIRO" in d or "ATM" in d:
        return "cash-withdrawal"
    if "SEGURO" in d:
        return "insurance"
    if "IMPUESTO" in d or "RETEFUENTE" in d:
        return "taxes"
    return None


# ─── SALARY XLSX ────────────────────────────────────────────────────────

def import_salary_xlsx(year: int) -> dict:
    """Import salary payment history from Reporte Salarios.xlsx."""
    xlsx_path = DECL_DIR / str(year) / "Reporte Salarios.xlsx"
    if not xlsx_path.exists():
        return {"source": "salary", "error": "File not found", "new": 0}

    print(f"  Reading: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["All Salary Payments"]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        date_str = str(data.get("Payment Date", ""))
        # Parse various date formats
        for fmt in ["%B %d, %Y", "%Y-%m-%d", "%b %d, %Y"]:
            try:
                # Handle "September Sep 27, 2024" format
                cleaned = re.sub(r"^\w+\s+", "", date_str) if re.match(r"\w+\s+\w+\s+\d", date_str) else date_str
                parsed = datetime.strptime(cleaned.strip(), fmt)
                date_str = parsed.strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

        amount_usd = float(data.get("Amount USD", 0) or 0)
        gross_cop = float(data.get("Gross COP (Official Rate)", 0) or 0)
        net_cop = float(data.get("Net COP Received (Est.)", 0) or 0)
        fx_loss = float(data.get("FX Loss COP (Deductible)", 0) or 0)
        rate = float(data.get("Exchange Rate (USD/COP)", 0) or 0)

        tid = tx_hash(date_str, amount_usd, f"salary-{data.get('Month', '')}", "salary")

        rec = {
            "id": tid,
            "date": date_str,
            "month": data.get("Month"),
            "month_name": data.get("Month Name"),
            "quarter": data.get("Quarter"),
            "amount_usd": amount_usd,
            "exchange_rate": rate,
            "gross_cop": round(gross_cop),
            "net_cop": round(net_cop),
            "fx_loss_cop": round(fx_loss),
            "income_type": str(data.get("Income Type", "")),
            "source": str(data.get("Source", "")),
            "payment_method": str(data.get("Payment Method", "")),
            "account": str(data.get("Account", "")),
            "tax_classification": str(data.get("Tax Classification", "")),
            "reference": str(data.get("Reference", "")),
            "uetr": str(data.get("UETR", "")),
        }
        records.append(rec)

    new_count = append_jsonl(SALARY_FILE, records)
    return {"source": "salary", "file": xlsx_path.name, "total": len(records), "new": new_count}


# ─── EXOGENA XLSX ──────────────────────────────────────────────────────

def import_exogena_xlsx(year: int) -> dict:
    """Import DIAN information exogena report."""
    xlsx_path = DECL_DIR / str(year) / "Reporte Informacion Exogena.xlsx"
    if not xlsx_path.exists():
        return {"source": "exogena", "error": "File not found", "new": 0}

    print(f"  Reading: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    records = []
    # Find the header row with NIT, Nombre, etc.
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "NIT" and row[1] and "Nombre" in str(row[1]):
            header_row = i
            headers = list(row)
            break
        # Also capture topes (thresholds)
        if row and row[4] and "Tope" in str(row[4]):
            tid = tx_hash(str(year), float(row[5] or 0), str(row[4]), "exogena-tope")
            records.append({
                "id": tid,
                "year": year,
                "type": "tope",
                "detail": str(row[4]),
                "value": float(row[5] or 0),
            })

    if header_row:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            data = dict(zip(headers, row))
            nit_reporter = str(data.get("NIT", ""))
            detail = str(data.get("Detalle", ""))
            value = float(data.get("Valor", 0) or 0)

            tid = tx_hash(f"{year}-{nit_reporter}", value, detail, "exogena")
            rec = {
                "id": tid,
                "year": year,
                "type": "third-party-report",
                "reporter_nit": nit_reporter,
                "reporter_name": str(data.get(headers[1], "")),
                "subject_nit": str(data.get(headers[2], "")),
                "subject_name": str(data.get(headers[3], "")),
                "detail": detail,
                "value": value,
                "tax_use": str(data.get(headers[6], "")),
                "additional_info": str(data.get(headers[7], "")),
            }
            records.append(rec)

    new_count = append_jsonl(EXOGENA_FILE, records)
    return {"source": "exogena", "file": xlsx_path.name, "total": len(records), "new": new_count}


# ─── E-INVOICES XLSX ───────────────────────────────────────────────────

def import_einvoices_xlsx(year: int) -> dict:
    """Import DIAN electronic invoices received report."""
    xlsx_path = DECL_DIR / str(year) / "Reporte Facturas Electronicas.xlsx"
    if not xlsx_path.exists():
        return {"source": "e-invoices", "error": "File not found", "new": 0}

    print(f"  Reading: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    # Find the header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] and "Identificación Emisor" in str(row[0]):
            header_row = i
            headers = [str(c) for c in row]
            break

    if not header_row:
        return {"source": "e-invoices", "error": "Header row not found", "new": 0}

    records = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[0]:
            continue

        emisor_nit = str(row[0])
        emisor_name = str(row[1] or "")
        fecha = str(row[2] or "")
        valor_facturado = float(row[3] or 0)
        valor_nc = float(row[4] or 0)
        valor_nd = float(row[5] or 0)
        valor_neto = float(row[6] or 0)
        valor_beneficio = float(row[7] or 0)
        medio_pago = str(row[8] or "")
        num_factura = str(row[9] or "")
        cufe = str(row[10] or "") if len(row) > 10 else ""

        tid = tx_hash(fecha, valor_facturado, f"{emisor_nit}-{num_factura}", "einvoice")

        rec = {
            "id": tid,
            "year": year,
            "date": fecha,
            "emisor_nit": emisor_nit,
            "emisor_name": emisor_name,
            "valor_facturado": valor_facturado,
            "valor_notas_credito": valor_nc,
            "valor_notas_debito": valor_nd,
            "valor_neto": valor_neto,
            "valor_beneficio_1pct": valor_beneficio,
            "medio_pago": medio_pago,
            "num_factura": num_factura,
            "cufe": cufe,
        }
        records.append(rec)

    new_count = append_jsonl(INVOICES_FILE, records)
    return {"source": "e-invoices", "file": xlsx_path.name, "total": len(records), "new": new_count}


# ─── 2023 TRANSFER CSVs ───────────────────────────────────────────────

def import_transfer_csvs(year: int) -> dict:
    """Import international/national transfer CSVs (email notification format)."""
    results = []

    for csv_name in ["international_transfers", "national_transfers"]:
        csv_path = DECL_DIR / str(year) / f"{csv_name}_{year}.csv"
        if not csv_path.exists():
            continue

        print(f"  Reading: {csv_path.name}")
        transactions = []

        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                fecha = row.get("Fecha", "").strip()
                # Parse YYYY/MM/DD
                try:
                    date_str = datetime.strptime(fecha, "%Y/%m/%d").strftime("%Y-%m-%d")
                except ValueError:
                    date_str = fecha

                valor_str = row.get("Valor Transacción", "0")
                valor = float(valor_str.replace(",", "").replace(".", "").strip() or "0")
                # Colombian format: dots are thousands, restore decimal
                if "," in row.get("Valor Transacción", ""):
                    # All values in these CSVs appear to be comma-separated thousands
                    valor = float(row.get("Valor Transacción", "0").replace(",", ""))

                description = row.get("Clase de Movimiento", "").strip().rstrip(",")
                lugar = row.get("Lugar de Transacción", "").strip()
                transfer_type = "international" if "international" in csv_name else "national"

                tid = tx_hash(date_str, valor, f"{description}-{lugar}", f"davivienda-{transfer_type}")

                tx = {
                    "id": tid,
                    "date": date_str,
                    "amount": valor,
                    "currency": "COP",
                    "description": f"{description} ({transfer_type})",
                    "category": "salary" if "Bancomex" in description else "transfer-in",
                    "bank": "davivienda",
                    "account": "davivienda-savings",
                    "raw": {
                        "lugar": lugar,
                        "transfer_type": transfer_type,
                        "source_csv": csv_path.name,
                    },
                }
                transactions.append(tx)

        new_count = append_jsonl(LEDGER_FILE, transactions)
        results.append({
            "source": f"{csv_name}",
            "file": csv_path.name,
            "total": len(transactions),
            "new": new_count,
        })

    return results if results else {"source": "transfer-csvs", "error": "No CSV files found", "new": 0}


# ─── MAIN ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import Declaracion data into finance-substrate")
    parser.add_argument("--year", type=int, default=2024, help="Tax year to import")
    parser.add_argument(
        "--source",
        choices=["all", "davivienda", "salary", "exogena", "einvoices", "transfers"],
        default="all",
        help="Which data source to import",
    )
    args = parser.parse_args()

    ensure_data_dirs()

    year_dir = DECL_DIR / str(args.year)
    if not year_dir.exists():
        print(f"Error: {year_dir} does not exist", file=sys.stderr)
        sys.exit(1)

    print(f"=== Importing Declaracion {args.year} ===\n")
    results = []

    importers = {
        "davivienda": import_davivienda_xlsx,
        "salary": import_salary_xlsx,
        "exogena": import_exogena_xlsx,
        "einvoices": import_einvoices_xlsx,
        "transfers": import_transfer_csvs,
    }

    sources = importers.keys() if args.source == "all" else [args.source]

    for source in sources:
        try:
            result = importers[source](args.year)
            if isinstance(result, list):
                results.extend(result)
            else:
                results.append(result)
        except Exception as e:
            results.append({"source": source, "error": str(e), "new": 0})
            print(f"  Error importing {source}: {e}")

    # Summary
    print(f"\n=== Import Summary ===")
    total_new = 0
    for r in results:
        status = f"{r.get('new', 0)} new" if "new" in r else r.get("error", "unknown")
        total_records = r.get("total", "")
        total_str = f" (of {total_records})" if total_records else ""
        print(f"  {r['source']}: {status}{total_str}")
        total_new += r.get("new", 0)

    print(f"\n  Total new records: {total_new}")
    print(f"  Data dir: {DATA_DIR}")

    return results


if __name__ == "__main__":
    main()
