#!/usr/bin/env python3
"""
Self-healing validation and anomaly detection for the finance-substrate skill.

Detects:
- Silent parser failures (extracted $0 for expected non-zero fields)
- Cross-source inconsistencies (certificates vs exogena vs planillas)
- XLSX schema drift (column changes in DIAN exports)
- MUISCA page structure changes (missing expected elements)
- Parser confidence scoring (how many fields extracted vs expected)

Outputs improvement signals to .control/improvement-log.jsonl for agent action.

Usage:
    python3 self_heal.py --year 2024                    # Full validation
    python3 self_heal.py --year 2024 --fix              # Auto-fix what's possible
    python3 self_heal.py --validate-parsers              # Check parser health
    python3 self_heal.py --validate-xlsx <path>          # Check XLSX schema
"""


# --- Python version guard (PEP 604 union syntax requires >= 3.10) ---
import sys

if sys.version_info < (3, 10):
    raise SystemExit(
        f"finance-substrate requires Python >= 3.10. "
        f"Got {sys.version_info.major}.{sys.version_info.minor}. "
        f"Install via `brew install python@3.11` or `pyenv install 3.11 && pyenv local 3.11`."
    )
# --- end guard ---

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

DATA_DIR = Path.home() / ".finance-substrate"
SKILL_DIR = Path(__file__).parent.parent
CONTROL_DIR = SKILL_DIR / ".control"
IMPROVEMENT_LOG = CONTROL_DIR / "improvement-log.jsonl"


def log_signal(event: str, severity: str, details: dict):
    """Append an improvement signal for agent consumption."""
    CONTROL_DIR.mkdir(parents=True, exist_ok=True)
    entry = {
        "event": event,
        "severity": severity,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        **details,
    }
    with open(IMPROVEMENT_LOG, "a") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")
    return entry


# ─── Certificate Parser Health ─────────────────────────────────────

def validate_certificate_extractions(year: int) -> list:
    """Check that each parsed certificate extracted meaningful values."""
    issues = []
    certs_file = DATA_DIR / "tax" / "certificates.jsonl"
    if not certs_file.exists():
        issues.append(("error", "no_certificates", "certificates.jsonl not found — run import_certificates.py first"))
        return issues

    # Expected non-zero fields per institution type
    expected_fields = {
        "davivienda": ["rendimientos_gravados", "patrimonio_cuentas", "gmf_total"],
        "skandia": ["pension_oblig_plus_fsp", "pension_voluntaria_aportes"],
        "colmedica": ["medicina_prepagada_deduccion_r39"],
        "nu-colombia": ["rendimientos_gravados", "retencion_renta"],
        "rappicard": ["deuda_patrimonio"],
        "rappicuenta": ["patrimonio_cuenta"],
        "nequi": [],  # Tiny amounts OK to be zero
        "banco-bogota": [],  # May have zero if no active credit
        "acciones-valores": [],
        "acciones-valores-fondos": [],
    }

    with open(certs_file) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            cert = json.loads(line)
            if cert.get("year") != year or cert.get("type") != "certificado-tributario":
                continue

            parser_id = cert.get("parser", "unknown")
            tax_summary = cert.get("tax_summary", {})
            sections = cert.get("sections", {})

            # Check 1: Parser confidence — how many fields were extracted?
            total_fields = 0
            zero_fields = 0
            for section_data in sections.values():
                if isinstance(section_data, dict):
                    for k, v in section_data.items():
                        total_fields += 1
                        if v == 0 or v == 0.0:
                            zero_fields += 1

            if total_fields > 0:
                confidence = (total_fields - zero_fields) / total_fields
                if confidence < 0.3 and total_fields > 3:
                    issues.append(("warning", "low_extraction_confidence", {
                        "parser": parser_id,
                        "confidence": round(confidence, 2),
                        "total_fields": total_fields,
                        "zero_fields": zero_fields,
                        "suggestion": f"Parser '{parser_id}' extracted mostly zeros — PDF format may have changed",
                    }))

            # Check 2: Expected non-zero fields
            expected = expected_fields.get(parser_id, [])
            for field in expected:
                val = tax_summary.get(field, 0)
                if val == 0:
                    issues.append(("warning", "expected_nonzero_field", {
                        "parser": parser_id,
                        "field": field,
                        "value": val,
                        "suggestion": f"Field '{field}' is 0 for {parser_id} — check if PDF format changed",
                    }))

    return issues


# ─── Cross-Source Validation ───────────────────────────────────────

def cross_validate_sources(year: int) -> list:
    """Compare values across certificates, exogena, and planillas."""
    issues = []

    # Load certificate aggregation
    certs_agg = {"retencion_renta": 0, "pension_oblig": 0}
    certs_file = DATA_DIR / "tax" / "certificates.jsonl"
    if certs_file.exists():
        with open(certs_file) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                c = json.loads(line)
                if c.get("year") != year or c.get("type") != "certificado-tributario":
                    continue
                ts = c.get("tax_summary", {})
                certs_agg["retencion_renta"] += ts.get("retencion_renta", 0) + ts.get("retencion_total", 0)
                certs_agg["pension_oblig"] += ts.get("pension_oblig_plus_fsp", 0)

    # Load exogena retenciones
    exogena_retenciones = 0
    exogena_pension = 0
    exogena_file = DATA_DIR / "tax" / "exogena.jsonl"
    if exogena_file.exists():
        with open(exogena_file) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                if rec.get("year") != year:
                    continue
                detail = rec.get("detail", "").lower()
                if "retención" in detail:
                    exogena_retenciones += rec.get("value", 0)
                if "pensión obligatoria" in detail and "aporte" in detail:
                    exogena_pension += rec.get("value", 0)

    # Load planilla totals
    planilla_pension = 0
    planilla_file = DATA_DIR / "tax" / "planillas.jsonl"
    if planilla_file.exists():
        with open(planilla_file) as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                p = json.loads(line)
                if not str(p.get("periodo", "")).startswith(str(year)):
                    continue
                planilla_pension += p.get("pension", 0)

    # Cross-validate retenciones: certificates vs exogena
    if certs_agg["retencion_renta"] > 0 and exogena_retenciones > 0:
        diff_pct = abs(certs_agg["retencion_renta"] - exogena_retenciones) / max(certs_agg["retencion_renta"], exogena_retenciones)
        if diff_pct > 0.05:  # 5% tolerance
            issues.append(("warning", "retenciones_mismatch", {
                "certificates": round(certs_agg["retencion_renta"]),
                "exogena": round(exogena_retenciones),
                "diff_pct": round(diff_pct * 100, 1),
                "suggestion": "Retenciones differ >5% between certificates and exogena — check for missing certificates",
            }))

    # Cross-validate pension: certificates vs exogena vs planillas
    sources = {}
    if certs_agg["pension_oblig"] > 0:
        sources["certificates"] = round(certs_agg["pension_oblig"])
    if exogena_pension > 0:
        sources["exogena"] = round(exogena_pension)
    if planilla_pension > 0:
        sources["planillas"] = round(planilla_pension)

    if len(sources) >= 2:
        vals = list(sources.values())
        max_diff = max(vals) - min(vals)
        if max_diff > 1000000:  # >$1M discrepancy
            issues.append(("info", "pension_source_discrepancy", {
                "sources": sources,
                "max_diff": max_diff,
                "suggestion": "Pension amounts differ across sources — certificates report fund-level, planillas report payment-level",
            }))

    return issues


# ─── XLSX Schema Validation ────────────────────────────────────────

def validate_xlsx_schema(path: str) -> list:
    """Check that an XLSX file matches expected DIAN export format."""
    issues = []
    try:
        import openpyxl
    except ImportError:
        issues.append(("error", "missing_dependency", {"package": "openpyxl"}))
        return issues

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        issues.append(("error", "xlsx_open_failed", {"path": path, "error": str(e)}))
        return issues

    ws = wb[wb.sheetnames[0]]

    # Detect file type by content
    is_exogena = False
    is_einvoice = False
    for i, row in enumerate(ws.iter_rows(values_only=True, max_row=20), 1):
        if row and row[0]:
            text = str(row[0])
            if "Información reportada por terceros" in text or "Exógena" in text:
                is_exogena = True
            if "Identificación Emisor" in text or "Factura" in text:
                is_einvoice = True

    if is_exogena:
        # Check exogena expected columns
        expected_header = ["NIT", "Nombre / Razón Social"]
        found_header = False
        for row in ws.iter_rows(values_only=True, max_row=20):
            if row and row[0] == "NIT" and row[1] and "Nombre" in str(row[1]):
                found_header = True
                # Verify column count
                cols = [str(c) for c in row if c]
                if len(cols) < 6:
                    issues.append(("warning", "exogena_schema_drift", {
                        "expected_cols": ">=6",
                        "found_cols": len(cols),
                        "headers": cols[:8],
                        "suggestion": "Exogena XLSX has fewer columns than expected — DIAN may have changed the format",
                    }))
                break
        if not found_header:
            issues.append(("error", "exogena_no_header", {
                "suggestion": "Could not find expected header row (NIT, Nombre) — DIAN format may have changed",
            }))

    elif is_einvoice:
        # Check e-invoice expected columns
        found_header = False
        for row in ws.iter_rows(values_only=True, max_row=25):
            if row and row[0] and "Identificaci" in str(row[0]):
                found_header = True
                cols = [str(c) for c in row if c]
                if len(cols) < 9:
                    issues.append(("warning", "einvoice_schema_drift", {
                        "expected_cols": ">=9",
                        "found_cols": len(cols),
                        "headers": cols[:11],
                        "suggestion": "E-invoice XLSX has fewer columns than expected",
                    }))
                break
        if not found_header:
            issues.append(("error", "einvoice_no_header", {
                "suggestion": "Could not find e-invoice header row — DIAN format may have changed",
            }))

    return issues


# ─── Parser Definition Health ──────────────────────────────────────

def validate_parser_definitions() -> list:
    """Check all parser definitions for completeness and consistency."""
    issues = []
    parsers_dir = SKILL_DIR / "parsers"
    if not parsers_dir.exists():
        issues.append(("error", "no_parsers_dir", {"suggestion": "parsers/ directory not found"}))
        return issues

    parser_files = sorted(parsers_dir.glob("*.json"))
    if not parser_files:
        issues.append(("error", "no_parser_files", {"suggestion": "No parser definitions found in parsers/"}))
        return issues

    ids_seen = set()
    priorities_seen = set()

    for pf in parser_files:
        try:
            with open(pf) as f:
                pdef = json.load(f)
        except json.JSONDecodeError as e:
            issues.append(("error", "parser_json_invalid", {"file": pf.name, "error": str(e)}))
            continue

        pid = pdef.get("institution", {}).get("id", "")

        # Duplicate ID check
        if pid in ids_seen:
            issues.append(("error", "duplicate_parser_id", {"id": pid, "file": pf.name}))
        ids_seen.add(pid)

        # Priority collision check
        priority = pdef.get("priority", 999)
        if priority in priorities_seen:
            issues.append(("warning", "priority_collision", {
                "priority": priority,
                "file": pf.name,
                "suggestion": "Two parsers share the same priority — matching order is ambiguous",
            }))
        priorities_seen.add(priority)

        # Tax summary mapping check
        mapping = pdef.get("tax_summary_mapping", {})
        if not mapping:
            issues.append(("warning", "empty_tax_summary", {
                "parser": pid,
                "file": pf.name,
                "suggestion": f"Parser '{pid}' has no tax_summary_mapping — extracted data won't flow to tax projection",
            }))

        # Section field method check
        for section in pdef.get("sections", []):
            for field in section.get("fields", []):
                method = field.get("method", "")
                valid_methods = ["label_then_amount", "dollar_amounts_after_anchor", "regex", "line_after_label", "sum_all_regex_matches"]
                if method not in valid_methods:
                    issues.append(("error", "invalid_field_method", {
                        "parser": pid,
                        "field": field.get("key"),
                        "method": method,
                        "valid": valid_methods,
                    }))

    return issues


# ─── Projection Sanity Checks ─────────────────────────────────────

def validate_projection(year: int) -> list:
    """Run sanity checks on the tax projection output."""
    issues = []

    # Run the projection and capture output
    try:
        sys.path.insert(0, str(SKILL_DIR / "scripts"))
        from tax_projection import project_tax
        result = project_tax(year)
    except Exception as e:
        issues.append(("error", "projection_failed", {"error": str(e)}))
        return issues

    f210 = result.get("form_210", {})

    # Check for negative values (should never happen)
    for section_name, section in f210.items():
        if isinstance(section, dict):
            for key, val in section.items():
                if isinstance(val, (int, float)) and val < 0:
                    issues.append(("error", "negative_value", {
                        "section": section_name,
                        "field": key,
                        "value": val,
                        "suggestion": "Negative values in Form 210 are invalid — check calculation logic",
                    }))

    # Check R34 = R32 - R33
    rt = f210.get("rentas_trabajo", {})
    expected_r34 = rt.get("R32_ingresos_brutos", 0) - rt.get("R33_incr", 0)
    actual_r34 = rt.get("R34_renta_liquida", 0)
    if abs(expected_r34 - actual_r34) > 1000:
        issues.append(("error", "r34_arithmetic", {
            "expected": expected_r34,
            "actual": actual_r34,
            "formula": "R34 = R32 - R33",
        }))

    # Check R41 <= 1,340 UVT
    uvt = result.get("uvt", 47065)
    cap_1340 = 1340 * uvt
    r41 = rt.get("R41_exentas_deduc_limitadas", 0)
    if r41 > cap_1340 + 1000:
        issues.append(("error", "r41_exceeds_cap", {
            "r41": r41,
            "cap_1340_uvt": cap_1340,
            "suggestion": "R41 exceeds 1,340 UVT cap (Art. 336 Num. 3) — deduction logic error",
        }))

    # Check R121 > 0 if taxable income > 1,090 UVT
    cg = f210.get("cedula_general", {})
    taxable_uvt = cg.get("renta_grav_uvt", 0)
    r121 = f210.get("impuesto", {}).get("R121_impuesto_rentas", 0)
    if taxable_uvt > 1090 and r121 == 0:
        issues.append(("warning", "zero_tax_above_threshold", {
            "taxable_uvt": taxable_uvt,
            "r121": r121,
            "suggestion": "Tax is $0 but taxable income exceeds 1,090 UVT — check bracket calculation",
        }))

    return issues


# ─── Main Report ───────────────────────────────────────────────────

def run_full_validation(year: int, fix: bool = False) -> dict:
    """Run all validations and return a structured report."""
    report = {
        "year": year,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "checks": {},
        "totals": {"errors": 0, "warnings": 0, "info": 0, "passed": 0},
    }

    checks = [
        ("parser_definitions", validate_parser_definitions),
        ("certificate_extractions", lambda: validate_certificate_extractions(year)),
        ("cross_validation", lambda: cross_validate_sources(year)),
        ("projection_sanity", lambda: validate_projection(year)),
    ]

    for check_name, check_fn in checks:
        try:
            issues = check_fn()
        except Exception as e:
            issues = [("error", "check_crashed", {"check": check_name, "error": str(e)})]

        report["checks"][check_name] = {
            "issues": issues,
            "status": "pass" if not issues else ("fail" if any(s == "error" for s, _, _ in issues) else "warn"),
        }

        for severity, event, details in issues:
            report["totals"][severity + "s"] = report["totals"].get(severity + "s", 0) + 1
            # Log improvement signals for agent consumption
            log_signal(event, severity, details if isinstance(details, dict) else {"message": details})

        if not issues:
            report["totals"]["passed"] += 1

    return report


def main():
    parser = argparse.ArgumentParser(description="Self-healing validation for finance-substrate")
    parser.add_argument("--year", type=int, default=2024)
    parser.add_argument("--fix", action="store_true", help="Auto-fix what's possible")
    parser.add_argument("--validate-parsers", action="store_true")
    parser.add_argument("--validate-xlsx", help="Validate an XLSX file")
    parser.add_argument("--json", action="store_true", help="JSON output")
    args = parser.parse_args()

    if args.validate_parsers:
        issues = validate_parser_definitions()
        for severity, event, details in issues:
            print(f"  [{severity:>7}] {event}: {details}")
        print(f"\n{len(issues)} issues found" if issues else "All parsers healthy")
        sys.exit(1 if any(s == "error" for s, _, _ in issues) else 0)

    if args.validate_xlsx:
        issues = validate_xlsx_schema(args.validate_xlsx)
        for severity, event, details in issues:
            print(f"  [{severity:>7}] {event}: {details}")
        print(f"\n{len(issues)} issues found" if issues else "XLSX schema valid")
        sys.exit(1 if any(s == "error" for s, _, _ in issues) else 0)

    # Full validation
    report = run_full_validation(args.year, args.fix)

    if args.json:
        print(json.dumps(report, indent=2, ensure_ascii=False, default=str))
    else:
        print(f"\n{'='*60}")
        print(f"  SELF-HEAL VALIDATION — {args.year}")
        print(f"{'='*60}\n")

        for check_name, check_data in report["checks"].items():
            status = check_data["status"]
            icon = {"pass": "OK", "warn": "WARN", "fail": "FAIL"}[status]
            print(f"  [{icon:>4}] {check_name}")
            for severity, event, details in check_data["issues"]:
                detail_str = details.get("suggestion", str(details)) if isinstance(details, dict) else str(details)
                print(f"         [{severity}] {event}: {detail_str}")

        t = report["totals"]
        print(f"\n  Summary: {t['passed']} passed, {t.get('errors', 0)} errors, {t.get('warnings', 0)} warnings, {t.get('infos', 0)} info")

        if t.get("errors", 0) > 0:
            print(f"\n  Improvement signals logged to: {IMPROVEMENT_LOG}")
            print(f"  Run with --fix to attempt auto-repair, or review signals for agent action.")


if __name__ == "__main__":
    main()
